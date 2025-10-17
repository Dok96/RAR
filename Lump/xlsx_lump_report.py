import time
from openpyxl import load_workbook
# from config import lump_report

# Функция для записи данных в файл .xlsx
def save_to_xlsx(fault_active, fault_type, length, fault_count,dist_lump_report):
    """
    Записывает данные о дефектах в файл .xlsx на основе шаблона.

    :param fault_active: Флаг активности дефекта (True/False).
    :param fault_type: Тип дефекта (например, "Lump", "Neckdown").
    :param length: Длина дефекта (в миллиметрах).
    :param lump_report: Путь к файлу шаблона .xlsx.
    """
    try:
        # Открываем существующий шаблон
        wb = load_workbook(dist_lump_report)
        ws = wb.active

       # инкремент для строк
        row = 19
        while ws[f"B{row}"]. value is not None:
            row +=1

        # Получаем текущее время
        timestamp = time.strftime("%Y-%m-%d %H:%M:%S")

        ##== обработка ономера ошибки
        FAULT_TYPES = {
            0: "no Error",
            1: "Lump",
            2: "Neckdown",
            3: "Lump and underexposed",
            4: "Neckdown and underexposed",
            5: "No optical power",
            6: "Device error"

        }
        fault_type = FAULT_TYPES.get(fault_type, "no Fault")


        # Записываем данные в ячейки, если флаг активен
        if fault_active:
            ws[f"B{row}"] = fault_type  # Тип дефекта
            ws[f"D{row}"] = length      # Длина дефекта
            ws[f"E{row}"] = timestamp   # Текущее время
            ws["B15"] = fault_count

        # Сохраняем изменения в файл
        wb.save(dist_lump_report)
        print(f"Данные сохранены в файл {dist_lump_report}.")

        return True

        # сбрасываем fault active
        #reset_fault_active()


    except FileNotFoundError:
        print(f"Ошибка: файл шаблона не найден по пути {dist_lump_report}.")
    except Exception as e:
        print(f"Ошибка при сохранении данных: {e}")

